from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# 1- Lê pasta de trabalho e planilha
wb = load_workbook("tabela/pivot_table.xlsx") #carrego o workbook e atribuo a var wb
sheet = wb["Relatório"] #passo o nome da planilha que quero trabalhar 

# 2- Referências das linhas e colunas (Python precisa sabe a coluna maxima e min, e as linhas tbm)
min_coluna = wb.active.min_column
max_coluna = wb.active.max_column
min_linha = wb.active.min_row
max_linha = wb.active.max_row

# 3- Adicionando Dados e Categorias no Gráfico
barchat = BarChart()

#tipo as legendas agora
data = Reference(
    sheet, #planilha
    min_col= min_coluna + 1, #min_coluna começa em ano
    max_col= max_coluna,
    min_row= min_linha,
    max_row= max_linha
)

categorias = Reference(
    sheet, #planilha
    min_col= min_coluna,
    max_col= min_coluna,
    min_row= min_linha + 1, #começa nos fabricantes e agora não quero eles
    max_row= max_linha
)

barchat.add_data(data, titles_from_data=True)
barchat.set_categories(categorias)

# 4- Criando o gráfico 
sheet.add_chart(barchat, "B10") #onde add
barchat.title = "Vendas por Fabricantes" #qual titulo
barchat.style = 2

# 5- Salvando o WorkBook
wb.save("tabela/barchart.xlsx")