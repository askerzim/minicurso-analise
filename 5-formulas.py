from openpyxl import load_workbook
from openpyxl.utils import get_column_letter #consegue pegar a letra da coluna

# 1 - Lê a pasta de trabalho e planilha
wb = load_workbook("tabela/barchart.xlsx")
sheet = wb["Relatório"]

# 2 - Referências das linhas e colunas (Python precisa sabe a coluna maxima e min, e as linhas tbm)
min_coluna = wb.active.min_column
max_coluna = wb.active.max_column
min_linha = wb.active.min_row
max_linha = wb.active.max_row

# 3 - Incluindo formulas
# sheet["B6"] = "SUM (B2:B5)"
# sheet["B6"].style = "Currency"
for i in range(min_coluna + 1, max_coluna + 1):
    letter = get_column_letter(i)
    sheet[f"{letter}{max_linha + 1}"] = f"=SUM({letter}{min_linha+1}:{letter}{max_linha})" #pra sempre add na linha 6
    sheet[f"{letter}{max_linha + 1}"].style  = "Currency" 
    
wb.save("final.xlsx")